"""
Appointment form + Ask AI app.
- Local: stores in data/appointments.xlsx
- Vercel: stores the same Excel file in Vercel Blob (persistent)
"""

from __future__ import annotations

import io
import json
import os
import re
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from flask import Flask, flash, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook

load_dotenv()

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
EXCEL_PATH = DATA_DIR / "appointments.xlsx"
BLOB_PATHNAME = "appointments.xlsx"
JSON_LEGACY_PATHNAME = "appointments.json"
COLUMNS = [
    "Timestamp",
    "BabyName",
    "Age",
    "Weight",
    "Date",
    "Time",
    "Area",
    "Mobile",
    "Email",
    "EditToken",
    "CalendarEventId",
]
ALLOWED_TIMES = ["12:00", "15:00", "18:00"]
TIME_LABELS = {
    "12:00": "12:00 PM · બપોરે ૧૨",
    "15:00": "3:00 PM · બપોરે ૩",
    "18:00": "6:00 PM · સાંજે ૬",
}
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

app = Flask(
    __name__,
    template_folder=str(APP_DIR / "templates"),
    static_folder=str(APP_DIR / "static"),
)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "appointment-local-dev-key")


class _VercelPathMiddleware:
    """Restore the browser path when Vercel rewrites to /api/index."""

    def __init__(self, wsgi_app):
        self.wsgi_app = wsgi_app

    def __call__(self, environ, start_response):
        original = (
            environ.get("HTTP_X_VERCEL_ORIGINAL_PATH")
            or environ.get("HTTP_X_INVOKE_PATH")
            or environ.get("HTTP_X_FORWARDED_URI")
            or ""
        )
        path = original.split("?", 1)[0]
        if path.startswith("http"):
            from urllib.parse import urlparse

            path = urlparse(path).path
        # If rewrite collapsed everything to /api/index, prefer the request URI path.
        req_uri = environ.get("REQUEST_URI") or environ.get("RAW_URI") or ""
        if (not path or path.startswith("/api/index")) and req_uri:
            path = req_uri.split("?", 1)[0]
        if path and path != environ.get("PATH_INFO"):
            environ["PATH_INFO"] = path
            environ["SCRIPT_NAME"] = ""
        return self.wsgi_app(environ, start_response)


app.wsgi_app = _VercelPathMiddleware(app.wsgi_app)

try:
    from werkzeug.middleware.proxy_fix import ProxyFix

    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
except Exception:
    pass


def using_blob() -> bool:
    return bool(os.getenv("BLOB_READ_WRITE_TOKEN", "").strip())


def storage_label() -> str:
    if using_blob():
        return "One shared file: appointments.xlsx (cloud)"
    return f"One shared file: {EXCEL_PATH}"


# --- Excel helpers -----------------------------------------------------------------

def _blank_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments"
    ws.append(COLUMNS)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rows_from_workbook_bytes(data: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    rows: list[dict[str, str]] = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map: dict[int, str] = {}
    for idx, h in enumerate(headers):
        name = str(h or "").strip()
        # Map legacy columns from the older form
        if name == "Name":
            name = "BabyName"
        elif name == "Reason":
            name = "Area"
        if name in COLUMNS:
            header_map[idx] = name
        elif idx < len(COLUMNS):
            header_map[idx] = COLUMNS[idx]

    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if excel_row is None or all(v is None or str(v).strip() == "" for v in excel_row):
            continue
        item: dict[str, str] = {c: "" for c in COLUMNS}
        for idx, value in enumerate(excel_row):
            key = header_map.get(idx)
            if key not in COLUMNS:
                continue
            if value is None:
                item[key] = ""
            elif isinstance(value, datetime):
                item[key] = (
                    value.strftime("%Y-%m-%d")
                    if key == "Date"
                    else value.strftime("%H:%M")
                    if key == "Time"
                    else value.strftime("%Y-%m-%d %H:%M:%S")
                )
            elif isinstance(value, time):
                item[key] = value.strftime("%H:%M")
            else:
                text = str(value).strip()
                if key == "Time" and re.fullmatch(r"\d{1,2}:\d{2}:\d{2}", text):
                    text = text[:5]
                item[key] = text
        if any(item[c] for c in ("BabyName", "Date", "Time", "Mobile", "EditToken")):
            rows.append(item)
    return rows


def _workbook_bytes_from_rows(rows: list[dict[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(c, "") for c in COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Blob storage ------------------------------------------------------------------

BLOB_API = "https://vercel.com/api/blob"


def _blob_list(prefix: str | None = None) -> list[dict[str, Any]]:
    import urllib.parse
    import urllib.request

    token = os.environ["BLOB_READ_WRITE_TOKEN"]
    qs = urllib.parse.urlencode({"prefix": prefix or BLOB_PATHNAME})
    req = urllib.request.Request(
        f"{BLOB_API}?{qs}",
        headers={
            "Authorization": f"Bearer {token}",
            "x-api-version": "7",
        },
        method="GET",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    return payload.get("blobs") or []


def _blob_download(url: str) -> bytes:
    import time as time_mod
    import urllib.parse
    import urllib.request

    # Bust CDN cache so freshly overwritten Excel is visible immediately.
    sep = "&" if "?" in url else "?"
    busted = f"{url}{sep}v={int(time_mod.time() * 1000)}"
    req = urllib.request.Request(
        busted,
        headers={"Cache-Control": "no-cache", "Pragma": "no-cache"},
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def _blob_upload(
    data: bytes,
    pathname: str = BLOB_PATHNAME,
    content_type: str = XLSX_CONTENT_TYPE,
) -> None:
    import urllib.error
    import urllib.parse
    import urllib.request

    token = os.environ["BLOB_READ_WRITE_TOKEN"]
    # Path-style PUT matches the Blob API contract used by x-api-version 7.
    url = f"{BLOB_API}/{urllib.parse.quote(pathname, safe='/')}"
    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": content_type,
            "x-api-version": "7",
            "x-vercel-blob-access": "public",
            "x-allow-overwrite": "1",
            "x-add-random-suffix": "0",
            "x-content-type": content_type,
            # Keep Excel reads fresh after each booking write.
            "x-cache-control-max-age": "0",
        },
        method="PUT",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            resp.read()
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Blob upload failed ({exc.code}): {detail}") from exc


def _find_blob(pathname: str) -> dict[str, Any] | None:
    blobs = _blob_list(pathname)
    return next((b for b in blobs if b.get("pathname") == pathname), None)


def _migrate_legacy_json_if_needed() -> bytes | None:
    """If an older JSON blob exists and Excel does not, convert once into Excel."""
    excel_blob = _find_blob(BLOB_PATHNAME)
    if excel_blob and excel_blob.get("url"):
        return None
    json_blob = _find_blob(JSON_LEGACY_PATHNAME)
    if not json_blob or not json_blob.get("url"):
        return None
    raw = _blob_download(json_blob["url"]).decode("utf-8")
    data = json.loads(raw or "[]")
    rows: list[dict[str, str]] = []
    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                rows.append({c: str(item.get(c, "") or "") for c in COLUMNS})
    excel_bytes = _workbook_bytes_from_rows(rows)
    _blob_upload(excel_bytes)
    return excel_bytes


def read_excel_bytes() -> bytes:
    """Return the single shared appointments.xlsx bytes (local or cloud)."""
    if using_blob():
        match = _find_blob(BLOB_PATHNAME)
        if match and match.get("url"):
            return _blob_download(match["url"])
        migrated = _migrate_legacy_json_if_needed()
        if migrated is not None:
            return migrated
        blank = _blank_workbook_bytes()
        _blob_upload(blank)
        return blank

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not EXCEL_PATH.exists():
        EXCEL_PATH.write_bytes(_blank_workbook_bytes())
    return EXCEL_PATH.read_bytes()


def write_excel_bytes(data: bytes) -> None:
    """Overwrite the same appointments.xlsx file (never create a new name)."""
    if using_blob():
        _blob_upload(data)
        return
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EXCEL_PATH.write_bytes(data)


def read_appointments() -> list[dict[str, str]]:
    return _rows_from_workbook_bytes(read_excel_bytes())


def find_booking(token: str) -> dict[str, str] | None:
    from flask import has_request_context, session

    token = (token or "").strip()
    if not token:
        return None
    for row in read_appointments():
        if row.get("EditToken") == token:
            return row
    # Fallback: serverless/CDN may lag right after write; use session copy.
    if has_request_context():
        cached = session.get(f"booking_{token}")
        if isinstance(cached, dict) and cached.get("EditToken") == token:
            return {c: str(cached.get(c, "") or "") for c in COLUMNS}
    return None


def remember_booking(booking: dict[str, str]) -> None:
    from flask import session

    token = booking.get("EditToken") or ""
    if not token:
        return
    session[f"booking_{token}"] = {c: booking.get(c, "") for c in COLUMNS}


def ensure_edit_tokens() -> list[dict[str, str]]:
    """Make sure every booking has an EditToken (for older rows)."""
    import uuid

    rows = read_appointments()
    changed = False
    for row in rows:
        if not row.get("EditToken"):
            row["EditToken"] = uuid.uuid4().hex
            changed = True
        for col in COLUMNS:
            row.setdefault(col, "")
    if changed:
        write_excel_bytes(_workbook_bytes_from_rows(rows))
    return rows


def save_booking_update(token: str, existing: dict[str, str]) -> tuple[bool, str | None]:
    """Validate POST and update booking. Returns (ok, error)."""
    booking, error = parse_booking_form(exclude_token=token)
    if error:
        return False, error
    assert booking is not None
    booking["EditToken"] = token
    booking["CalendarEventId"] = existing.get("CalendarEventId", "")
    try:
        from google_calendar import update_khatna_event

        event_id = update_khatna_event(booking.get("CalendarEventId", ""), booking)
        if event_id:
            booking["CalendarEventId"] = event_id
    except Exception:
        pass
    if not update_appointment(token, booking):
        return False, "Could not update this booking."
    remember_booking(booking)
    return True, None


def append_appointment(booking: dict[str, str]) -> None:
    """Append one row into the same Excel workbook and save it back."""
    data = read_excel_bytes()
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    # If this is a legacy workbook, rewrite headers to the new schema once.
    headers = [str(ws.cell(1, i).value or "") for i in range(1, ws.max_column + 1)]
    if "BabyName" not in headers or "EditToken" not in headers:
        existing = _rows_from_workbook_bytes(data)
        write_excel_bytes(_workbook_bytes_from_rows(existing + [booking]))
        return

    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        ws.append(COLUMNS)
    ws.append([booking.get(c, "") for c in COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    write_excel_bytes(buf.getvalue())


def update_appointment(token: str, booking: dict[str, str]) -> bool:
    rows = read_appointments()
    found = False
    for i, row in enumerate(rows):
        if row.get("EditToken") == token:
            booking["EditToken"] = token
            booking["Timestamp"] = row.get("Timestamp") or booking.get("Timestamp", "")
            if not booking.get("CalendarEventId"):
                booking["CalendarEventId"] = row.get("CalendarEventId", "")
            rows[i] = {c: booking.get(c, "") for c in COLUMNS}
            found = True
            break
    if not found:
        return False
    write_excel_bytes(_workbook_bytes_from_rows(rows))
    return True


def delete_appointment(token: str) -> dict[str, str] | None:
    """Remove a booking. Returns the deleted row (or None)."""
    rows = read_appointments()
    kept: list[dict[str, str]] = []
    deleted: dict[str, str] | None = None
    for row in rows:
        if row.get("EditToken") == token and deleted is None:
            deleted = row
            continue
        kept.append(row)
    if deleted is None:
        return None
    write_excel_bytes(_workbook_bytes_from_rows(kept))
    return deleted


def parse_booking_form(exclude_token: str | None = None) -> tuple[dict[str, str] | None, str | None]:
    """Validate form POST. Returns (booking, error_message)."""
    baby_name = request.form.get("baby_name", "").strip()
    age = request.form.get("age", "").strip()
    weight = request.form.get("weight", "").strip()
    date_str = request.form.get("date", "").strip()
    time_str = request.form.get("time", "").strip()
    area = request.form.get("area", "").strip()
    mobile = request.form.get("mobile", "").strip()
    email = request.form.get("email", "").strip()

    if not all([baby_name, age, weight, date_str, time_str, area, mobile]):
        return None, "Please fill in all required fields. · કૃપા કરીને બધા જરૂરી ખાના ભરો."
    if time_str not in ALLOWED_TIMES:
        return None, "Please choose 12 PM, 3 PM, or 6 PM only. · માત્ર ૧૨, ૩ અથવા ૬ વાગ્યાનો સમય પસંદ કરો."
    if not re.fullmatch(r"[0-9+\-\s]{8,15}", mobile):
        return None, "Enter a valid mobile number. · યોગ્ય મોબાઇલ નંબર લખો."

    return (
        {
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "BabyName": baby_name,
            "Age": age,
            "Weight": weight,
            "Date": date_str,
            "Time": time_str,
            "Area": area,
            "Mobile": mobile,
            "Email": email,
            "EditToken": "",
            "CalendarEventId": "",
        },
        None,
    )


# --- Ask / filter ------------------------------------------------------------------

STOP_WORDS = {
    "tell", "show", "list", "what", "when", "have", "appointment", "appointments",
    "please", "find", "give", "with", "from", "that", "this", "my", "the", "and",
    "are", "who", "booked", "schedule", "schedules", "me", "any", "get", "see",
    "display", "looking", "look", "for", "can", "you", "your", "our", "there",
    "here", "about", "all", "everything", "everyone", "rows", "data", "excel",
    "do", "i", "is", "a", "an", "of", "on", "to", "in", "at", "or", "be", "was",
    "were", "am", "pm", "o", "clock", "time", "date", "name", "reason", "which",
    "ones", "those", "these", "some", "them", "their", "has", "had", "will",
    "how", "many", "booking", "bookings", "khatna", "patient", "patients",
    "afternoon", "evening", "morning", "noon", "today", "tomorrow", "mobile",
    "number", "phone", "baby", "area",
}


def _normalize_time_token(raw: str) -> time | None:
    raw = raw.strip().lower().replace(".", ":")
    raw = re.sub(r"\s+", "", raw)
    raw = raw.replace("o'clock", "").replace("oclock", "")

    m = re.fullmatch(r"(\d{1,2}):(\d{2})\s*(am|pm)?", raw)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2))
        meridiem = m.group(3)
        if meridiem == "pm" and hour != 12:
            hour += 12
        if meridiem == "am" and hour == 12:
            hour = 0
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return time(hour, minute)
        return None

    m = re.fullmatch(r"(\d{1,2})\s*(am|pm)", raw)
    if m:
        hour = int(m.group(1))
        meridiem = m.group(2)
        if meridiem == "pm" and hour != 12:
            hour += 12
        if meridiem == "am" and hour == 12:
            hour = 0
        if 0 <= hour <= 23:
            return time(hour, 0)
        return None

    # Bare hour: "6", "18"
    m = re.fullmatch(r"(\d{1,2})", raw)
    if m:
        hour = int(m.group(1))
        if 0 <= hour <= 23:
            return time(hour, 0)
        return None

    return None


def _row_time(value: str) -> time | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p", "%I %p", "%I%p"):
        try:
            return datetime.strptime(text, fmt).time().replace(second=0, microsecond=0)
        except ValueError:
            continue
    return _normalize_time_token(text)


def _row_date(value: str) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    try:
        # Lenient parse for values like "Aug 16, 2026"
        parsed = datetime.strptime(text, "%b %d, %Y")
        return parsed.replace(hour=0, minute=0, second=0, microsecond=0)
    except ValueError:
        return None


def _extract_time_queries(q: str) -> list[tuple[time, bool, bool]]:
    """
    Returns list of (time, ambiguous_ampm, hour_only).
    hour_only=True means match any minutes in that hour.
    """
    found: list[tuple[time, bool, bool]] = []

    patterns = [
        # at 6:00 / 6:00pm / 18:00
        r"(?:at|around|about|by|for)?\s*(\d{1,2}:\d{2})\s*(am|pm)?",
        # 6pm / 6 pm
        r"(?:at|around|about|by|for)?\s*(\d{1,2})\s*(am|pm)\b",
        # at 6 / at 6 o'clock
        r"(?:at|around|about|by)\s+(\d{1,2})(?:\s*o'?clock)?\b",
        # bare evening-style: "6 o'clock"
        r"\b(\d{1,2})\s*o'?clock\b",
    ]

    seen: set[tuple[int, int, bool, bool]] = set()
    for pattern in patterns:
        for m in re.finditer(pattern, q, flags=re.IGNORECASE):
            groups = [g for g in m.groups() if g is not None]
            if not groups:
                continue
            token = "".join(groups)
            # Rebuild nicer token
            if len(groups) == 2 and groups[1] in {"am", "pm"}:
                token = f"{groups[0]}{groups[1]}"
            elif len(groups) == 1:
                token = groups[0]
            t = _normalize_time_token(token)
            if not t:
                continue
            cleaned = re.sub(r"\s+", "", token.lower())
            has_meridiem = cleaned.endswith(("am", "pm"))
            hour_only = ":" not in cleaned and not has_meridiem
            # "6pm" is hour_only for minutes (any minute in that hour) but not ambiguous am/pm
            if has_meridiem and ":" not in cleaned:
                hour_only = True
            ambiguous = not has_meridiem
            key = (t.hour, t.minute, ambiguous, hour_only)
            if key in seen:
                continue
            seen.add(key)
            found.append((t, ambiguous, hour_only))
    return found


def _time_matches_row(row_t: time, target: time, ambiguous: bool, hour_only: bool) -> bool:
    if hour_only:
        if ambiguous and 1 <= target.hour <= 12:
            return row_t.hour % 12 == target.hour % 12
        return row_t.hour == target.hour

    if ambiguous and 1 <= target.hour <= 12:
        return row_t.hour % 12 == target.hour % 12 and row_t.minute == target.minute
    return row_t.hour == target.hour and row_t.minute == target.minute


def _fuzzy_contains(haystack: str, needle: str) -> bool:
    h = haystack.lower().strip()
    n = needle.lower().strip()
    if not n:
        return False
    if n in h:
        return True
    # Soft: all needle tokens appear somewhere
    parts = [p for p in re.split(r"[\s_-]+", n) if p]
    return bool(parts) and all(p in h for p in parts)


def _human_time(value: str) -> str:
    plain = {"12:00": "12:00 PM", "15:00": "3:00 PM", "18:00": "6:00 PM"}
    return plain.get(value, TIME_LABELS.get(value, value or "—"))


def _format_booking_line(row: dict[str, str]) -> str:
    return (
        f"{row.get('BabyName') or '—'} | {_human_time(row.get('Time', ''))} on {row.get('Date') or '—'} | "
        f"Age {row.get('Age') or '—'}, Weight {row.get('Weight') or '—'} | "
        f"{row.get('Area') or '—'} | Mobile {row.get('Mobile') or '—'}"
        + (f" | {row.get('Email')}" if row.get("Email") else "")
    )


def _doctor_summary(question: str, matched: list[dict[str, str]], notes: list[str] | None = None) -> str:
    q = question.strip()
    if not matched:
        return f'No bookings matched "{q}". Try another time, date, baby name, area, or mobile.'

    lines = [f'Found {len(matched)} booking(s) for "{q}":']
    for row in matched[:12]:
        lines.append(f"• {_format_booking_line(row)}")
    if len(matched) > 12:
        lines.append(f"• …and {len(matched) - 12} more.")
    if notes:
        lines.append("Filters used: " + "; ".join(notes) + ".")
    return "\n".join(lines)


def filter_with_rules(question: str, rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], str]:
    q = question.strip().lower()
    if not rows:
        return [], "No appointments saved yet."

    if re.search(
        r"\b(all|everything|everyone|entire|full\s+list|show\s+me\s+all|list\s+all)\b",
        q,
    ) and not re.search(r"\b(\d{1,2}|today|tomorrow|named|about|afternoon|evening|morning)\b", q):
        return rows, _doctor_summary(question, rows)

    filtered = list(rows)
    notes: list[str] = []

    if re.search(r"\b(afternoon|noon)\b", q) and not _extract_time_queries(q):
        filtered = [r for r in filtered if r.get("Time") in {"12:00", "15:00"}]
        notes.append("afternoon slots (12 PM / 3 PM)")
    elif re.search(r"\bevening\b", q) and not _extract_time_queries(q):
        filtered = [r for r in filtered if r.get("Time") == "18:00"]
        notes.append("evening slot (6 PM)")
    elif re.search(r"\bmorning\b", q) and not _extract_time_queries(q):
        filtered = [r for r in filtered if r.get("Time") == "12:00"]
        notes.append("noon slot (12 PM)")

    time_queries = _extract_time_queries(q)
    if time_queries:
        def matches_time(val: str) -> bool:
            row_t = _row_time(val)
            if row_t is None:
                return False
            return any(
                _time_matches_row(row_t, t, ambiguous, hour_only)
                for t, ambiguous, hour_only in time_queries
            )

        filtered = [r for r in filtered if matches_time(r.get("Time", ""))]
        pretty_bits = []
        for t, ambiguous, hour_only in time_queries:
            label = t.strftime("%I").lstrip("0")
            if not hour_only:
                label += t.strftime(":%M")
            if ambiguous:
                label += " (am/pm)"
            else:
                label += " " + t.strftime("%p")
            pretty_bits.append(label)
        notes.append("time " + ", ".join(pretty_bits))

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if re.search(r"\btoday\b", q):
        filtered = [r for r in filtered if _row_date(r.get("Date", "")) == today]
        notes.append("today")
    elif re.search(r"\btomorrow\b", q):
        tomorrow = today + timedelta(days=1)
        filtered = [r for r in filtered if _row_date(r.get("Date", "")) == tomorrow]
        notes.append("tomorrow")
    else:
        date_match = re.search(r"(\d{4}-\d{2}-\d{2})", q)
        if date_match:
            target = datetime.strptime(date_match.group(1), "%Y-%m-%d")
            filtered = [r for r in filtered if _row_date(r.get("Date", "")) == target]
            notes.append(f"date {date_match.group(1)}")

    mobile_match = re.search(r"(\d{8,15})", q)
    if mobile_match:
        digits = re.sub(r"\D", "", mobile_match.group(1))
        filtered = [r for r in filtered if digits in re.sub(r"\D", "", r.get("Mobile", ""))]
        notes.append(f"mobile {digits}")

    name_match = re.search(
        r"(?:for|named|name(?:\s+is)?|with|by|baby)\s+([a-zA-Z][a-zA-Z\s'-]{0,40})",
        q,
    )
    if name_match:
        name = name_match.group(1).strip()
        name = re.split(
            r"\b(?:at|on|about|with|tomorrow|today|around|appointment|appointments)\b",
            name,
        )[0].strip()
        if name and name not in STOP_WORDS:
            filtered = [r for r in filtered if _fuzzy_contains(r.get("BabyName", ""), name)]
            notes.append(f'baby name "{name}"')

    about_match = re.search(r"\b(?:about|regarding|area|from|in)\s+([a-zA-Z0-9][\w\s'-]{0,40})", q)
    if about_match:
        keyword = about_match.group(1).strip()
        keyword = re.split(r"\b(?:at|on|named|with|tomorrow|today)\b", keyword)[0].strip()
        if keyword and keyword not in STOP_WORDS and keyword not in {"appointment", "appointments", "khatna"}:
            area_hits = [r for r in filtered if _fuzzy_contains(r.get("Area", ""), keyword)]
            name_hits = [r for r in filtered if _fuzzy_contains(r.get("BabyName", ""), keyword)]
            if area_hits:
                filtered = area_hits
                notes.append(f'area "{keyword}"')
            elif name_hits:
                filtered = name_hits
                notes.append(f'baby name "{keyword}"')
            elif "about" in q or "area" in q or "regarding" in q or " in " in f" {q} ":
                filtered = area_hits
                notes.append(f'area "{keyword}"')

    if not notes:
        words = [w for w in re.findall(r"[a-zA-Z0-9]{2,}", q) if w not in STOP_WORDS]
        words = [w for w in words if not w.isdigit()]
        if words:
            def row_hits(r: dict[str, str]) -> bool:
                blob = " ".join(
                    r.get(c, "")
                    for c in ("BabyName", "Age", "Weight", "Date", "Time", "Area", "Mobile", "Email")
                ).lower()
                return any(w in blob for w in words)

            hits = [r for r in filtered if row_hits(r)]
            if hits:
                filtered = hits
                notes.append("text " + ", ".join(words))
            else:
                return [], _doctor_summary(question, [])

    if notes and not filtered:
        return [], _doctor_summary(question, [])

    return filtered, _doctor_summary(question, filtered, notes)


def ask_free_ai(question: str, rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], str] | None:
    if not rows:
        return None

    today = datetime.now().strftime("%Y-%m-%d")
    tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    catalog = []
    for i, row in enumerate(rows):
        catalog.append({
            "index": i,
            "baby_name": row.get("BabyName", ""),
            "age": row.get("Age", ""),
            "weight": row.get("Weight", ""),
            "date": row.get("Date", ""),
            "time": row.get("Time", ""),
            "time_label": _human_time(row.get("Time", "")),
            "area": row.get("Area", ""),
            "mobile": row.get("Mobile", ""),
            "email": row.get("Email", ""),
        })

    system = (
        "You are a careful clinic assistant for a doctor managing khatna (circumcision) bookings. "
        "Available appointment times are only 12:00 (12 PM), 15:00 (3 PM), and 18:00 (6 PM). "
        "Interpret casual language: 'at 6' / '6pm' / 'evening' => 18:00; "
        "'at 3' / '3pm' may mean 15:00; 'noon' / '12' / 'afternoon' may mean 12:00 or 15:00. "
        "Use the provided today/tomorrow dates. "
        "Return ONLY valid JSON with keys: indices (matching catalog index values) and summary. "
        "In summary: say how many matches, then list each with baby name, date, time, area, and mobile. "
        "If nothing matches, indices=[] and suggest what to try next. "
        "Be precise. Do not invent bookings."
    )
    user = (
        f"Today's date: {today}\n"
        f"Tomorrow's date: {tomorrow}\n"
        f"Doctor question: {question}\n\n"
        f"Booking catalog as JSON:\n{json.dumps(catalog, ensure_ascii=False)}"
    )

    providers: list[tuple[str, str, str]] = []
    groq_key = os.getenv("GROQ_API_KEY", "").strip()
    if groq_key:
        providers.append((groq_key, "https://api.groq.com/openai/v1", "llama-3.3-70b-versatile"))
        providers.append((groq_key, "https://api.groq.com/openai/v1", "llama-3.1-8b-instant"))
    openai_key = os.getenv("OPENAI_API_KEY", "").strip()
    if openai_key:
        providers.append((openai_key, "https://api.openai.com/v1", "gpt-4o-mini"))

    for api_key, base_url, model in providers:
        try:
            from openai import OpenAI

            client = OpenAI(api_key=api_key, base_url=base_url)
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            payload = json.loads(response.choices[0].message.content or "{}")
            indices = payload.get("indices") or []
            summary = str(payload.get("summary") or "").strip()
            valid = [i for i in indices if isinstance(i, int) and 0 <= i < len(rows)]
            matched = [rows[i] for i in valid]
            if not summary:
                summary = _doctor_summary(question, matched)
            elif matched and summary.count("•") < min(2, len(matched)):
                extra = "\n".join(f"• {_format_booking_line(r)}" for r in matched)
                if extra not in summary:
                    summary = summary.rstrip() + "\n" + extra
            return matched, summary
        except Exception:
            continue
    return None



def doctor_password() -> str:
    return os.getenv("DOCTOR_PASSWORD", "doctor123").strip() or "doctor123"


def google_calendar_link() -> str:
    """Open Google Calendar; deep-link to configured calendar when possible."""
    from urllib.parse import quote

    cal_id = os.getenv("GOOGLE_CALENDAR_ID", "").strip()
    if cal_id:
        return f"https://calendar.google.com/calendar/u/0/r?cid={quote(cal_id)}"
    return "https://calendar.google.com/calendar/"


def doctor_logged_in() -> bool:
    from flask import session

    return bool(session.get("doctor_ok"))


def require_doctor(view):
    from functools import wraps

    from flask import session

    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("doctor_ok"):
            return redirect(url_for("doctor_login"))
        return view(*args, **kwargs)

    return wrapped


@app.route("/", methods=["GET", "POST"])
@app.route("/book", methods=["GET", "POST"])
@app.route("/api/index", methods=["GET", "POST"])
def patient_form():
    """Public patient link — khatna booking form only."""
    import uuid

    if request.method == "POST":
        booking, error = parse_booking_form()
        if error:
            flash(error, "error")
            return redirect(url_for("patient_form"))

        assert booking is not None
        booking["EditToken"] = uuid.uuid4().hex
        try:
            try:
                from google_calendar import create_khatna_event

                event_id = create_khatna_event(booking)
                if event_id:
                    booking["CalendarEventId"] = event_id
            except Exception:
                pass
            append_appointment(booking)
            remember_booking(booking)
            return redirect(url_for("booking_confirmation", token=booking["EditToken"]))
        except Exception as exc:
            flash(f"Could not submit right now. Please try again. ({exc})", "error")
            return redirect(url_for("patient_form"))

    return render_template(
        "patient.html",
        allowed_times=ALLOWED_TIMES,
        time_labels=TIME_LABELS,
        booking=None,
        edit_mode=False,
        doctor_mode=False,
        form_action=url_for("patient_form"),
        cancel_url=None,
    )


@app.route("/confirmation/<token>")
def booking_confirmation(token: str):
    booking = find_booking(token)
    if not booking:
        flash("Booking not found. You can submit a new appointment below.", "error")
        return redirect(url_for("patient_form"))
    return render_template(
        "confirmation.html",
        booking=booking,
        time_labels=TIME_LABELS,
        edit_url=url_for("edit_booking", token=token),
    )


@app.route("/edit/<token>", methods=["GET", "POST"])
def edit_booking(token: str):
    existing = find_booking(token)
    if not existing:
        flash("Booking not found or link expired. Please submit a new appointment.", "error")
        return redirect(url_for("patient_form"))

    if request.method == "POST":
        ok, error = save_booking_update(token, existing)
        if not ok:
            flash(error or "Could not update this booking.", "error")
            return redirect(url_for("edit_booking", token=token))
        flash("Booking updated successfully. · બુકિંગ અપડેટ થઈ ગયું.", "success")
        return redirect(url_for("booking_confirmation", token=token))

    return render_template(
        "patient.html",
        allowed_times=ALLOWED_TIMES,
        time_labels=TIME_LABELS,
        booking=existing,
        edit_mode=True,
        doctor_mode=False,
        form_action=url_for("edit_booking", token=token),
        cancel_url=url_for("booking_confirmation", token=token),
    )


@app.route("/doctor/edit/<token>", methods=["GET", "POST"])
@require_doctor
def doctor_edit_booking(token: str):
    ensure_edit_tokens()
    existing = find_booking(token)
    if not existing:
        flash("Booking not found.", "error")
        return redirect(url_for("doctor_dashboard"))

    if request.method == "POST":
        ok, error = save_booking_update(token, existing)
        if not ok:
            flash(error or "Could not update this booking.", "error")
            return redirect(url_for("doctor_edit_booking", token=token))
        flash("Booking updated.", "success")
        return redirect(url_for("doctor_dashboard"))

    return render_template(
        "patient.html",
        allowed_times=ALLOWED_TIMES,
        time_labels=TIME_LABELS,
        booking=existing,
        edit_mode=True,
        doctor_mode=True,
        form_action=url_for("doctor_edit_booking", token=token),
        cancel_url=url_for("doctor_dashboard"),
    )


@app.route("/doctor/clear/<token>", methods=["POST"])
@require_doctor
def doctor_clear_booking(token: str):
    ensure_edit_tokens()
    deleted = delete_appointment(token)
    if not deleted:
        flash("Booking not found.", "error")
        return redirect(url_for("doctor_dashboard"))

    try:
        from google_calendar import delete_khatna_event

        delete_khatna_event(deleted.get("CalendarEventId", ""))
    except Exception:
        pass

    # Drop cleared rows from the open AI chat thread
    from flask import session

    history = list(session.get("doctor_chat") or [])
    cleaned = []
    for turn in history:
        results = [r for r in (turn.get("results") or []) if r.get("EditToken") != token]
        turn = dict(turn)
        turn["results"] = results
        cleaned.append(turn)
    session["doctor_chat"] = cleaned

    name = deleted.get("BabyName") or "Booking"
    flash(f"Cleared booking for {name}.", "success")
    return redirect(url_for("doctor_dashboard"))


@app.route("/doctor/login", methods=["GET", "POST"])
def doctor_login():
    from flask import session

    if session.get("doctor_ok"):
        return redirect(url_for("doctor_dashboard"))

    if request.method == "POST":
        password = request.form.get("password", "")
        if password == doctor_password():
            session["doctor_ok"] = True
            flash("Welcome back, doctor.", "success")
            return redirect(url_for("doctor_dashboard"))
        flash("Incorrect password.", "error")

    return render_template("doctor_login.html")


@app.route("/doctor/logout", methods=["POST", "GET"])
def doctor_logout():
    from flask import session

    session.clear()
    flash("Signed out.", "success")
    return redirect(url_for("doctor_login"))


@app.route("/doctor", methods=["GET", "POST"])
@require_doctor
def doctor_dashboard():
    from flask import session

    ai_status = (
        "AI assistant connected"
        if os.getenv("GROQ_API_KEY", "").strip() or os.getenv("OPENAI_API_KEY", "").strip()
        else "Smart filter active"
    )
    try:
        from google_calendar import calendar_configured

        calendar_status = (
            "Google Calendar connected"
            if calendar_configured()
            else "Google Calendar not connected yet"
        )
    except Exception:
        calendar_status = "Google Calendar not connected yet"
    chat_history = list(session.get("doctor_chat") or [])
    # Ensure tokens exist before ask/list so Edit links always work.
    ensure_edit_tokens()

    if request.method == "POST":
        action = request.form.get("action")

        if action == "clear_chat":
            session["doctor_chat"] = []
            chat_history = []
            flash("Chat cleared. Ask a new question anytime.", "success")
            return redirect(url_for("doctor_dashboard"))

        if action == "ask":
            question = request.form.get("question", "").strip()
            rows = read_appointments()
            if not question:
                flash("Type a question first.", "error")
            else:
                ai = ask_free_ai(question, rows)
                if ai is not None:
                    ask_result, ask_summary = ai
                else:
                    ask_result, ask_summary = filter_with_rules(question, rows)
                chat_history.append(
                    {
                        "question": question,
                        "summary": ask_summary,
                        "results": ask_result,
                    }
                )
                # Keep the thread manageable in the session cookie
                chat_history = chat_history[-12:]
                session["doctor_chat"] = chat_history
            return redirect(url_for("doctor_dashboard"))

    appointments = ensure_edit_tokens()

    def sort_key(row: dict[str, str]):
        return (row.get("Date") or "", row.get("Time") or "", row.get("Timestamp") or "")

    appointments = sorted(appointments, key=sort_key)
    return render_template(
        "doctor.html",
        appointments=appointments,
        booking_count=len(appointments),
        chat_history=chat_history,
        excel_path=storage_label(),
        ai_status=ai_status,
        calendar_status=calendar_status,
        calendar_link=google_calendar_link(),
        patient_link=url_for("patient_form", _external=True),
        time_labels=TIME_LABELS,
    )


@app.route("/download.xlsx")
@require_doctor
def download_excel():
    data = read_excel_bytes()
    return send_file(
        io.BytesIO(data),
        mimetype=XLSX_CONTENT_TYPE,
        as_attachment=True,
        download_name="appointments.xlsx",
    )


# WSGI entry for Vercel
app.debug = False

if __name__ == "__main__":
    print(f"Storage: {storage_label()}")
    print("Patient form: http://127.0.0.1:5000/")
    print("Doctor desk:  http://127.0.0.1:5000/doctor")
    app.run(debug=True, host="127.0.0.1", port=5000)
